from openpyxl import Workbook, load_workbook, workbook


def change_the_square(workbook, letter, number, value):
    square = letter + str(number)
    ws = workbook.active
    ws[square].value = value
    return

def save_excel_file(workbook, path):
    workbook.save(path)


def fill_the_name_column(workbook, letter, values):
    '''values je v obliki seznama'''
    i=0
    while i < len(values):
        change_the_square(workbook, letter, i+1, values[i])
        i+=1
    return

def set_column_width(path, letter):
    wb = load_workbook(path)
    ws = wb.active
    ws.column_dimensions[letter].width = 25
    save_excel_file(wb, path)


def fill_the_attribute_column(workbook, letter, values, dictionary_of_values):
    '''values je v obliki seznama'''
    probability_list = []

    change_the_square(workbook, letter, 1, values[0])
    i=1
    while i < len(values):

        probability = dictionary_of_values[values[i]] / len(values[1:])
        probability_list.append(probability)
        change_the_square(workbook, letter, i+1, values[i] + "            " + "{:.2f}".format(probability * 100) + "%")
        i+=1

    return probability_list

def fill_the_object_rarity_column(workbook, letter, nft_probability_list):
    '''values je v obliki seznama'''
    change_the_square(workbook, letter, 1, "probability")

    i=0
    l = len(nft_probability_list)
    while i < len(nft_probability_list[0]):
        product = 1
        for j in range(l):
            product *= nft_probability_list[j][i] 

        percentage = "{:.8f}".format(product * 100) + " %"
        
        change_the_square(workbook, letter, i+2, percentage)
        i+=1
    return





        







